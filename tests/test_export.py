import numpy as np
from openpyxl import load_workbook

from measurer.export import export_measured_batch
from measurer.image_queue import ImageQueue, RectRoi
from measurer.measurement import measure_image
from measurer.synthetic import SingleMetalIslandSpec, create_single_metal_island_image


def test_export_blocks_when_there_are_no_measured_images(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    image_path = source_folder / "pending.tif"
    queue = ImageQueue()
    queue.add_image_data(image_path, np.ones((8, 8), dtype=np.uint8))

    result = export_measured_batch(queue)

    assert result.exported_count == 0
    assert result.blocked_reason == "No measured images to export."
    assert result.message == "No measured images to export."
    assert not (source_folder / "measured_image").exists()
    assert not (source_folder / "debug_image").exists()
    assert queue.rows[0].export_status == "Not exported"


def test_export_writes_single_source_artifacts_for_measured_images(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    image_path = source_folder / "stem_zc.tif"
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(image_path, image)
    queue.record_measurement_result(0, measure_image(image, roi=None))

    result = export_measured_batch(queue)

    measured_folder = source_folder / "measured_image"
    debug_folder = source_folder / "debug_image"
    assert result.exported_count == 1
    assert result.output_folder == source_folder
    assert result.message == "Exported 1 measured image."
    assert (measured_folder / "stem_zc_result.png").is_file()
    assert (debug_folder / "stem_zc_debug.png").is_file()
    assert (measured_folder / "measurements.xlsx").is_file()
    assert queue.rows[0].export_status == "Exported"


def test_export_requires_output_folder_for_multi_source_measured_images(tmp_path):
    first_folder = tmp_path / "first"
    second_folder = tmp_path / "second"
    first_folder.mkdir()
    second_folder.mkdir()
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(first_folder / "first.tif", image)
    queue.add_image_data(second_folder / "second.tif", image)
    queue.record_measurement_result(0, measure_image(image, roi=None))
    queue.record_measurement_result(1, measure_image(image, roi=None))

    result = export_measured_batch(queue)

    assert result.needs_output_folder is True
    assert result.exported_count == 0
    assert result.message == "Choose an output folder for multi-source export."
    assert not (first_folder / "measured_image").exists()
    assert not (second_folder / "debug_image").exists()
    assert queue.rows[0].export_status == "Not exported"
    assert queue.rows[1].export_status == "Not exported"


def test_export_writes_multi_source_artifacts_under_chosen_output_folder(tmp_path):
    first_folder = tmp_path / "first"
    second_folder = tmp_path / "second"
    output_folder = tmp_path / "export"
    first_folder.mkdir()
    second_folder.mkdir()
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(first_folder / "first.tif", image)
    queue.add_image_data(second_folder / "second.tif", image)
    queue.record_measurement_result(0, measure_image(image, roi=None))
    queue.record_measurement_result(1, measure_image(image, roi=None))

    result = export_measured_batch(queue, output_folder=output_folder)

    measured_folder = output_folder / "measured_image"
    debug_folder = output_folder / "debug_image"
    assert result.exported_count == 2
    assert result.output_folder == output_folder
    assert result.message == "Exported 2 measured images."
    assert (measured_folder / "first_result.png").is_file()
    assert (measured_folder / "second_result.png").is_file()
    assert (debug_folder / "first_debug.png").is_file()
    assert (debug_folder / "second_debug.png").is_file()
    assert (measured_folder / "measurements.xlsx").is_file()
    assert not (first_folder / "measured_image").exists()
    assert not (second_folder / "debug_image").exists()
    assert queue.rows[0].export_status == "Exported"
    assert queue.rows[1].export_status == "Exported"


def test_export_multi_source_same_filename_uses_mvs_overwrite_rule(tmp_path):
    first_folder = tmp_path / "first"
    second_folder = tmp_path / "second"
    output_folder = tmp_path / "export"
    first_folder.mkdir()
    second_folder.mkdir()
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(first_folder / "same_name.tif", image)
    queue.add_image_data(second_folder / "same_name.tif", image)
    queue.record_measurement_result(0, measure_image(image, roi=None))
    queue.record_measurement_result(1, measure_image(image, roi=None))

    result = export_measured_batch(queue, output_folder=output_folder)

    measured_files = sorted(
        path.name for path in (output_folder / "measured_image").iterdir()
    )
    debug_files = sorted(path.name for path in (output_folder / "debug_image").iterdir())
    assert result.exported_count == 2
    assert measured_files == ["measurements.xlsx", "same_name_result.png"]
    assert debug_files == ["same_name_debug.png"]
    assert queue.rows[0].export_status == "Exported"
    assert queue.rows[1].export_status == "Exported"


def test_export_requires_overwrite_confirmation_before_replacing_existing_targets(
    tmp_path,
):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    image_path = source_folder / "stem_zc.tif"
    measured_folder = source_folder / "measured_image"
    debug_folder = source_folder / "debug_image"
    measured_folder.mkdir()
    debug_folder.mkdir()
    existing_result = measured_folder / "stem_zc_result.png"
    existing_debug = debug_folder / "stem_zc_debug.png"
    existing_workbook = measured_folder / "measurements.xlsx"
    existing_result.write_text("old result")
    existing_debug.write_text("old debug")
    existing_workbook.write_text("old workbook")
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(image_path, image)
    queue.record_measurement_result(0, measure_image(image, roi=None))

    result = export_measured_batch(queue)

    assert result.overwrite_required is True
    assert result.exported_count == 0
    assert result.message == "Confirm overwrite to export."
    assert result.overwrite_summary is not None
    assert result.overwrite_summary.output_folder == source_folder
    assert result.overwrite_summary.result_image_count == 1
    assert result.overwrite_summary.debug_image_count == 1
    assert result.overwrite_summary.workbook_count == 1
    assert existing_result.read_text() == "old result"
    assert existing_debug.read_text() == "old debug"
    assert existing_workbook.read_text() == "old workbook"
    assert queue.rows[0].export_status == "Not exported"


def test_export_overwrites_existing_targets_when_confirmed(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    image_path = source_folder / "stem_zc.tif"
    measured_folder = source_folder / "measured_image"
    debug_folder = source_folder / "debug_image"
    measured_folder.mkdir()
    debug_folder.mkdir()
    existing_result = measured_folder / "stem_zc_result.png"
    existing_debug = debug_folder / "stem_zc_debug.png"
    existing_workbook = measured_folder / "measurements.xlsx"
    existing_result.write_bytes(b"old result")
    existing_debug.write_bytes(b"old debug")
    existing_workbook.write_bytes(b"old workbook")
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(image_path, image)
    queue.record_measurement_result(0, measure_image(image, roi=None))

    result = export_measured_batch(queue, overwrite_existing=True)

    assert result.overwrite_required is False
    assert result.exported_count == 1
    assert result.message == "Exported 1 measured image."
    assert existing_result.read_bytes() != b"old result"
    assert existing_debug.read_bytes() != b"old debug"
    assert existing_workbook.read_bytes() != b"old workbook"
    assert queue.rows[0].export_status == "Exported"


def test_export_includes_only_measured_images_in_files_and_workbook(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    measured_path = source_folder / "measured.tif"
    pending_path = source_folder / "pending.tif"
    failed_path = source_folder / "failed.tif"
    measured_image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    failed_image = np.ones((128, 128), dtype=np.uint8) * 20
    queue = ImageQueue()
    queue.add_image_data(measured_path, measured_image)
    queue.add_image_data(pending_path, measured_image)
    queue.add_image_data(failed_path, failed_image)
    queue.set_group([0], "Process A")
    queue.record_measurement_result(0, measure_image(measured_image, roi=None))
    queue.record_measurement_failure(2, measure_image(failed_image, roi=None))

    result = export_measured_batch(queue)

    measured_folder = source_folder / "measured_image"
    debug_folder = source_folder / "debug_image"
    assert result.message == "Exported 1 measured image. Skipped 1 pending, 1 failed."
    assert (measured_folder / "measured_result.png").is_file()
    assert (debug_folder / "measured_debug.png").is_file()
    assert not (measured_folder / "pending_result.png").exists()
    assert not (debug_folder / "failed_debug.png").exists()
    assert queue.rows[0].export_status == "Exported"
    assert queue.rows[1].export_status == "Not exported"
    assert queue.rows[2].export_status == "Not exported"

    workbook = load_workbook(measured_folder / "measurements.xlsx")
    measurement_rows = list(workbook["Measurements"].iter_rows(values_only=True))
    trace_rows = list(workbook["Trace"].iter_rows(values_only=True))
    assert {row[0] for row in measurement_rows[1:]} == {"measured.tif"}
    assert {row[0] for row in trace_rows[1:]} == {"measured.tif"}
    assert all(row[1] == "Process A" for row in measurement_rows[1:])


def test_export_summary_separates_units_and_uses_current_scale(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    first_path = source_folder / "scaled.tif"
    second_path = source_folder / "pixel.tif"
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    queue = ImageQueue()
    queue.add_image_data(first_path, image)
    queue.add_image_data(second_path, image)
    queue.set_group([0, 1], "Process A")
    queue.set_manual_scale(0, "0.5")
    queue.record_measurement_result(0, measure_image(image, roi=None))
    queue.record_measurement_result(1, measure_image(image, roi=None))

    export_measured_batch(queue)

    workbook = load_workbook(source_folder / "measured_image" / "measurements.xlsx")
    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))[1:]
    tcd_rows = [
        row for row in summary_rows if row[0] == "Process A" and row[1] == "TCD"
    ]
    assert sorted((row[3], row[8]) for row in tcd_rows) == [(16.0, "nm"), (32.0, "px")]

    measurement_rows = list(workbook["Measurements"].iter_rows(values_only=True))[1:]
    tcd_measurements = [
        row for row in measurement_rows if row[2] == "TCD"
    ]
    assert sorted((row[0], row[5], row[6]) for row in tcd_measurements) == [
        ("pixel.tif", 32.0, "px"),
        ("scaled.tif", 16.0, "nm"),
    ]


def test_export_trace_records_custom_roi_geometry(tmp_path):
    source_folder = tmp_path / "source"
    source_folder.mkdir()
    image_path = source_folder / "roi.tif"
    image = create_single_metal_island_image(
        SingleMetalIslandSpec(
            image_width=128,
            image_height=128,
            center_x=64,
            top_y=24,
            height=60,
            tcd=32,
            bcd=48,
        )
    )
    roi = RectRoi(x=32, y=16, width=64, height=80)
    queue = ImageQueue()
    queue.add_image_data(image_path, image)
    queue.set_roi(0, roi)
    queue.record_measurement_result(0, measure_image(image, roi=roi))

    export_measured_batch(queue)

    workbook = load_workbook(source_folder / "measured_image" / "measurements.xlsx")
    trace_rows = list(workbook["Trace"].iter_rows(values_only=True))
    first_trace = trace_rows[1]
    assert first_trace[13:18] == ("rectangle", 32, 16, 64, 80)
