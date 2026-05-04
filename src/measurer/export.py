from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from statistics import mean, median, pstdev

import numpy as np
from openpyxl import Workbook
from PySide6.QtWidgets import QApplication

from measurer.debug_render import debug_panel_qimage
from measurer.image_queue import ImageQueue
from measurer.measurement import MeasurementResult
from measurer.measurement_report import (
    MEASUREMENT_TYPE_ORDER,
    measurement_report_key,
)
from measurer.result_render import result_qimage
from measurer.trace_sheet import TRACE_SHEET_HEADER, trace_sheet_row


NO_MEASURED_IMAGES_MESSAGE = "No measured images to export."
MULTI_SOURCE_OUTPUT_FOLDER_MESSAGE = "Choose an output folder for multi-source export."
OVERWRITE_CONFIRMATION_MESSAGE = "Confirm overwrite to export."
_QAPPLICATION: QApplication | None = None


@dataclass(frozen=True)
class OverwriteSummary:
    output_folder: Path
    result_image_count: int = 0
    debug_image_count: int = 0
    workbook_count: int = 0

    @property
    def has_existing_targets(self) -> bool:
        return (
            self.result_image_count > 0
            or self.debug_image_count > 0
            or self.workbook_count > 0
        )


@dataclass(frozen=True)
class ExportResult:
    exported_count: int
    skipped_pending_count: int = 0
    skipped_failed_count: int = 0
    output_folder: Path | None = None
    blocked_reason: str = ""
    message: str = ""
    needs_output_folder: bool = False
    overwrite_required: bool = False
    overwrite_summary: OverwriteSummary | None = None


def export_measured_batch(
    queue: ImageQueue,
    output_folder: Path | None = None,
    overwrite_existing: bool = False,
) -> ExportResult:
    measured_rows = [
        (row_index, row)
        for row_index, row in enumerate(queue.rows)
        if row.measure_status == "Measured"
        and isinstance(row.measurement_results, MeasurementResult)
    ]
    if not measured_rows:
        return ExportResult(
            exported_count=0,
            blocked_reason=NO_MEASURED_IMAGES_MESSAGE,
            message=NO_MEASURED_IMAGES_MESSAGE,
        )

    source_folders = {row.path.parent for _, row in measured_rows}
    if output_folder is None and len(source_folders) != 1:
        return ExportResult(
            exported_count=0,
            blocked_reason=MULTI_SOURCE_OUTPUT_FOLDER_MESSAGE,
            message=MULTI_SOURCE_OUTPUT_FOLDER_MESSAGE,
            needs_output_folder=True,
        )

    if output_folder is None:
        output_folder = next(iter(source_folders))

    measured_folder = output_folder / "measured_image"
    debug_folder = output_folder / "debug_image"
    planned_rows = [
        (
            row_index,
            row,
            measured_folder / f"{row.path.stem}_result.png",
            debug_folder / f"{row.path.stem}_debug.png",
        )
        for row_index, row in measured_rows
    ]
    workbook_path = measured_folder / "measurements.xlsx"
    overwrite_summary = _existing_target_summary(
        output_folder=output_folder,
        result_paths=[result_path for _, _, result_path, _ in planned_rows],
        debug_paths=[debug_path for _, _, _, debug_path in planned_rows],
        workbook_path=workbook_path,
    )
    if overwrite_summary.has_existing_targets and not overwrite_existing:
        return ExportResult(
            exported_count=0,
            output_folder=output_folder,
            blocked_reason=OVERWRITE_CONFIRMATION_MESSAGE,
            message=OVERWRITE_CONFIRMATION_MESSAGE,
            overwrite_required=True,
            overwrite_summary=overwrite_summary,
        )

    measured_folder.mkdir(parents=True, exist_ok=True)
    debug_folder.mkdir(parents=True, exist_ok=True)

    for row_index, row, result_path, debug_path in planned_rows:
        scale = queue.resolve_scale(row_index)
        result = row.measurement_results
        _save_result_image(
            row.image,
            result,
            scale.nm_per_px,
            result_path,
        )
        _save_debug_image(row.image, result, debug_path)

    _write_workbook(queue, workbook_path)
    queue.record_export_success([row_index for row_index, _ in measured_rows])

    skipped_pending_count = sum(
        1 for row in queue.rows if row.measure_status == "Pending"
    )
    skipped_failed_count = sum(1 for row in queue.rows if row.measure_status == "Failed")
    exported_count = len(measured_rows)
    return ExportResult(
        exported_count=exported_count,
        skipped_pending_count=skipped_pending_count,
        skipped_failed_count=skipped_failed_count,
        output_folder=output_folder,
        message=_format_export_message(
            exported_count, skipped_pending_count, skipped_failed_count
        ),
    )


def _existing_target_summary(
    output_folder: Path,
    result_paths: list[Path],
    debug_paths: list[Path],
    workbook_path: Path,
) -> OverwriteSummary:
    return OverwriteSummary(
        output_folder=output_folder,
        result_image_count=sum(1 for path in set(result_paths) if path.exists()),
        debug_image_count=sum(1 for path in set(debug_paths) if path.exists()),
        workbook_count=1 if workbook_path.exists() else 0,
    )


def _save_result_image(
    image: np.ndarray,
    result: MeasurementResult,
    nm_per_px: float | None,
    output_path: Path,
) -> None:
    _ensure_qapplication()
    result_qimage(image, result, nm_per_px).save(str(output_path))


def _save_debug_image(
    image: np.ndarray, result: MeasurementResult, output_path: Path
) -> None:
    _ensure_qapplication()
    debug_panel_qimage(image, result).save(str(output_path))


def _write_workbook(queue: ImageQueue, output_path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    measurements_sheet = workbook.create_sheet("Measurements")
    trace_sheet = workbook.create_sheet("Trace")

    summary_sheet.append(
        [
            "group",
            "measurement type",
            "count",
            "mean",
            "median",
            "min",
            "max",
            "std",
            "unit",
        ]
    )
    measurements_sheet.append(
        ["file", "group", "measurement type", "target ID", "status", "value", "unit"]
    )
    trace_sheet.append(TRACE_SHEET_HEADER)

    summary_values: dict[tuple[str, str, str], list[float]] = {}
    for row_index, row in enumerate(queue.rows):
        if row.measure_status != "Measured":
            continue
        if not isinstance(row.measurement_results, MeasurementResult):
            continue

        result = row.measurement_results
        scale_resolution = queue.resolve_scale(row_index)
        unit = "px" if scale_resolution.nm_per_px is None else "nm"
        scale = 1.0 if scale_resolution.nm_per_px is None else scale_resolution.nm_per_px
        for measurement in result.measurements.values():
            report_key = measurement_report_key(measurement)
            if report_key is None:
                continue
            measurement_type = report_key.measurement_type
            target_id = report_key.target_id
            value = (
                None
                if measurement.status != "success"
                else round(measurement.value_px * scale, 1)
            )
            measurements_sheet.append(
                [
                    row.file_name,
                    row.group,
                    measurement_type,
                    target_id,
                    measurement.status,
                    value,
                    unit,
                ]
            )
            if measurement.status == "success" and value is not None:
                summary_values.setdefault((row.group, measurement_type, unit), []).append(
                    value
                )
            trace_sheet.append(
                trace_sheet_row(
                    file_name=row.file_name,
                    group=row.group,
                    measurement=measurement,
                    measurement_type=measurement_type,
                    target_id=target_id,
                    result=result,
                    scale_source=scale_resolution.source,
                    scale_nm_per_px=scale_resolution.nm_per_px,
                    roi=row.roi,
                )
            )

    for (group, measurement_type, unit), values in sorted(
        summary_values.items(),
        key=lambda item: (
            item[0][0],
            MEASUREMENT_TYPE_ORDER.index(item[0][1]),
            item[0][2],
        ),
    ):
        summary_sheet.append(
            [
                group,
                measurement_type,
                len(values),
                round(mean(values), 1),
                round(median(values), 1),
                round(min(values), 1),
                round(max(values), 1),
                round(pstdev(values), 1),
                unit,
            ]
        )
    workbook.save(output_path)


def _format_export_message(
    exported_count: int, skipped_pending_count: int, skipped_failed_count: int
) -> str:
    image_label = "image" if exported_count == 1 else "images"
    message = f"Exported {exported_count} measured {image_label}."
    skipped_parts = []
    if skipped_pending_count:
        skipped_parts.append(f"{skipped_pending_count} pending")
    if skipped_failed_count:
        skipped_parts.append(f"{skipped_failed_count} failed")
    if skipped_parts:
        message += f" Skipped {', '.join(skipped_parts)}."
    return message


def _ensure_qapplication() -> None:
    global _QAPPLICATION
    _QAPPLICATION = QApplication.instance() or QApplication([])
